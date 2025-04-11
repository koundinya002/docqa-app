## Importing necessary libraries
import os
import sys
import base64
from dotenv import load_dotenv
# file manipulation
import pdfplumber
import fitz
from PIL import Image
from pathlib import Path
from openpyxl import Workbook, load_workbook
#openAI
import openai
from openai import OpenAI
#langchain
from langchain.chat_models import ChatOpenAI
from langchain.prompts import ChatPromptTemplate
from langchain.schema.output_parser import StrOutputParser
from langchain.schema.messages import HumanMessage, AIMessage
from langchain.schema import Document
from langchain_text_splitters import CharacterTextSplitter
#Chromadb
import chromadb
from chromadb.utils import embedding_functions
 
## 
# ANSI escape sequences for colors
RED = '\033[91m'
GREEN = '\033[92m'
YELLOW = '\033[93m'
BLUE = '\033[94m'
MAGENTA = '\033[95m'
CYAN = '\033[96m'
WHITE = '\033[97m'
RESET = '\033[0m' 
 
## GLOBAL VARIABLES
# set up input, output and database directories
 
input_files_path = "input/" 
output_images_path = "output"
db_path = 'dbpath'
excel_path = 'excel'
if not os.path.exists(output_images_path):
    os.makedirs(output_images_path)
if not os.path.exists(db_path):
    os.makedirs(db_path)
if not os.path.exists(excel_path):
    os.makedirs(excel_path)
 
# set up openAI api key and create a client
# load environment variables
load_dotenv() 
    ## uncomment the following to set the api key directly without dotenv
    #os.environ["OPENAI_API_KEY"] ="YOUR-OPENAI-API-KEY"
    #api_key = os.environ["OPENAI_API_KEY"]
api_key = ''
try:
    api_key = os.environ["OPENAI_API_KEY"]
except KeyError:
    print(RED+'Please set up the OpenAI API. Exitting ...'+RESET)
    sys.exit()
client = OpenAI(api_key=api_key)

# use embedding model
openai_ef = embedding_functions.OpenAIEmbeddingFunction(api_key=api_key, model_name="text-embedding-3-small")
# use LLM models
chain_gpt_35 = ChatOpenAI(model="gpt-3.5-turbo",api_key = api_key, max_tokens=1024)
chain_gpt_4_vision = ChatOpenAI(model="gpt-4o",  api_key = api_key,max_tokens=1024)
 
# set up Chroma. The parameter `path` must be a local path on the machine where Chroma is running.
chroma_client = chromadb.PersistentClient(path=db_path)
collection_name="docqa_chroma_collection"
text_image_collection=chroma_client.get_or_create_collection(name=collection_name, embedding_function=openai_ef)
 
def extract_text_from_pdf(pdf_path):
    '''
    Extract texts from input pdf and stores it into chromaDB

    Args:
        pdf_path: path of a single pdf file
    return:
        None
    '''
    with pdfplumber.open(pdf_path) as pdf:
        pdf_name = os.path.basename(pdf_path)
        #print(pdf_name)
        for page_num, page in enumerate(pdf.pages):
        text = page.extract_text()
        #print(f"Page {page_num +1} Text:\n{text}\n")
        docs = [Document(page_content=text, metadata={"source": pdf_name,"pageno":page_num + 1,"type":"text"})]
        text_splitter = CharacterTextSplitter(
            separator="\n\n",
            chunk_size=100,
            chunk_overlap=20,
            length_function=len,
            is_separator_regex=False,
            )
        
        texts = text_splitter.split_documents(docs)
        
        for i in range(len(texts)):
            # Add the documents to the Chroma collection
            text_image_collection.add(
            ids=[pdf_path + "__" + str(texts[i].metadata["pageno"]) + "__" + str(i)],
            documents=[texts[i].page_content],
            metadatas=[texts[i].metadata])
 
def extract_images_from_pdf(pdf_path, output_folder):
    '''
    Extract images from input pdf and stores it into chromaDB after encoding

    Args:
        pdf_path: path of a single pdf file
        output_folder: path where extracted images are stored
    return:
        None
    '''
    document = fitz.open(pdf_path)
    pdf_name = os.path.basename(pdf_path)
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)
    for page_num in range(document.page_count):
        page = document[page_num]
        images = page.get_images(full=True)
        for img_index, img in enumerate(images):
        xref = img[0]
        base_image = document.extract_image(xref)
        image_bytes = base_image["image"]
        image_ext = base_image["ext"]
        image_filename = f"{output_folder}/{pdf_name}_page_{page_num +1}_image_{img_index+1}.{image_ext}"
    
        with open(image_filename, "wb") as image_file:
            image_file.write(image_bytes)
            
        #print(f"Saved image: {image_filename}")
        with open(image_filename, 'rb') as image_file:
            encoded_image = base64.b64encode(image_file.read()).decode('utf-8')
        image_summary = summarize_image(encoded_image)
            
        docs = [Document(page_content=image_summary, metadata={"source": pdf_name,"pageno":page_num +1,"type":"image","image_path":image_filename})]
        text_splitter = CharacterTextSplitter(
            separator="\n\n",
            chunk_size=1000,
            chunk_overlap=20,
            length_function=len,
            is_separator_regex=False,
            )
        texts = text_splitter.split_documents(docs)
        for i in range(len(texts)):
            # Add the documents to the collection
            text_image_collection.add(
            ids=[pdf_path + "__" + str(texts[i].metadata["pageno"]) + "__" + str(i)],
            documents=[texts[i].page_content],
            metadatas=[texts[i].metadata])
    
def summarize_image(encoded_image):
    '''
    summarize an image by invoking LLM

    Args:
        encoded_image: encoded image object
    return:
        summary content
    '''
    prompt = [
        AIMessage(content="You are a bot that is good at analyzing images."),
        HumanMessage(content=[
            {"type": "text", "text": "Describe the contents of this image."},
            {
                "type": "image_url",
                "image_url": {
                    "url": f"data:image/png;base64,{encoded_image}"
                },
            },
        ])
    ]
    try:
        response = chain_gpt_4_vision.invoke(prompt)
    except :
        prompt = [
            AIMessage(content="You are a bot that is good at analyzing images."),
            HumanMessage(content=[
                {"type": "text", "text": "Describe the contents of this image."},
                {
                    "type": "image_url",
                    "image_url": {
                        "url": f"data:image/png;base64,{encoded_image}"
                    },
                },
            ])
        ]
    return response.content
 
def extract_all_from_pdf(pdf_path, output_folder):
    '''
    a wrapper function of extract_text_from_pdf and extract_images_from_pdf
    '''
    #print("Extracting text...")
    extract_text_from_pdf(pdf_path)
    #print("Extracting images...")
    extract_images_from_pdf(pdf_path, output_folder)
 
def encode_image(image_path):
    """
    Encode an image file to base64 format.

    Args:
        image_path (str): Path to the image file.

    Returns:
        str: Base64 encoded string representation of the image.
    """
    # Open the image
    image = Image.open(image_path)
 
    # Check if the format is supported
    if image.format.lower() not in ['png', 'jpeg', 'gif', 'webp']:
        # Convert the image to 'jpeg'
        byte_arr = io.BytesIO()
        image.save(byte_arr, format='JPEG')
        encoded_image = base64.b64encode(byte_arr.getvalue()).decode('utf-8')
    else:
        # Encode the image in base64
        with open(image_path, 'rb') as image_file:
            encoded_image = base64.b64encode(image_file.read()).decode('utf-8')
 
    return encoded_image
 
def context_extracter(similar_doc):
    """
    Extract context and encoded images from metadata of a document.

    Args:
        similar_doc (dict): Dictionary containing metadata and documents.

    Returns:
        tuple: A tuple containing:
            - context (str): Concatenated text context from 'text' metadata.
            - list_encoded_images (list): List of base64 encoded images.
            - model (str): Updated model name based on image encoding.
    """
    Model = "gpt-3.5-turbo"
    list_encoded_images = []
    context = ""
    encoded_image_paths = set()  # Set to track already encoded image paths
 
    for i in range(len(similar_doc["metadatas"][0])):
        metadata = similar_doc["metadatas"][0][i]
        
        # Check if the type is image and within the limit
        if metadata["type"] == "image" and i <= 5:
            image_path = metadata["image_path"]
            # Only encode if the image path has not been encoded yet
            if image_path not in encoded_image_paths:
                list_encoded_images.append(encode_image(image_path))
                encoded_image_paths.add(image_path)  # Mark this image as encoded
                Model = "gpt-4o"  # Update model if an image is encoded
 
        # Check if the type is text
        if metadata["type"] == "text":
            context += similar_doc["documents"][0][i] + "\n"
    
    return context, list_encoded_images, Model
       
def model_response(context, image_encodings, model_name, question):
    """
    Use llm to get the answer to a question

    Args:
        context: Concatenated text context from 'text' metadata.
        image_encodings: image contents returned form db querying
        model_name: selected llm model
        question: question string

    Returns:
        model_response: the llm response
    """
    # Define the text context
    text_context = f"(question)\n{question} \n (reference context){context}"
 
    # Define the model response for "gpt-4o"
    if model_name == "gpt-4o":
        # Define the image context
        image_context = []
        for image_encoding in image_encodings:
            image_context.append({"type":"image_url","image_url":{"url":f"data:image/png;base64,{image_encoding}"}})
 
        # Create the model response
        model_response = client.chat.completions.create(
            model=model_name,
            messages=[
                {"role":"system","content":"You are a helpful assistant that responds in markdown. Help me in answering the question."},
                {"role":"user","content":[{"type":"text","text":text_context}] + image_context}
            ],
            temperature=0.0,
        )
        return model_response
 
    # Define the model response for "gpt-3.5-turbo"
    elif model_name == "gpt-3.5-turbo":
        # Create the model response
        model_response = client.chat.completions.create(
            model=model_name,
            messages=[
                {"role":"system","content":"You are a helpful assistant that responds in markdown. Help me in answering the question."},
                {"role":"user","content":[{"type":"text","text":text_context}]}
            ],
            temperature=0.0,
        )
        return model_response
 
def ask_question(question):
    """
    Entry function. Perform db query by providing the question string and get the relevant context 

    Args: 
        question: the question string
    Returns:
        a tuple containing:
        -response: answer to the question
        -relevant_pdf: the relevant pdf where the answer is found
        -relevant_page: page number
    """
    # qa_pairs.append((question, response))
    similar_doc = text_image_collection.query(query_texts=question, n_results=5)
    # ! uncomment below to see relevant documents 
    # print("========\n metadata",similar_doc["metadatas"][0],"\n===========")
    context,image_encodings,ModelName = context_extracter(similar_doc)
    response = model_response(context = context,image_encodings= image_encodings, question= question ,model_name= ModelName)
    relevant_doc = similar_doc["metadatas"][0]
    relevant_pdf = relevant_doc[0]["source"]
    relevant_page = relevant_doc[0]["pageno"]
    print("PDF: ", relevant_pdf)
    print("Page No. : ", relevant_page)
    return (response, relevant_pdf, relevant_page)
 
def save_qa_to_excel(data):
    """
    Export the output to an excel file
    Args: 
        data: the output data
    Returns:
        None
    """
    # Specify the file path to save the Excel file
    excel_file = 'excel/output.xlsx'
 
    if not Path(excel_file).is_file():
        workbook = Workbook()
    else:
        workbook = load_workbook(excel_file)
        
    sheet = workbook.active
    max_row = sheet.max_row
 
    # Header row
    sheet['A1'] = 'Question'
    sheet['B1'] = 'Answer'
    sheet['C1'] = 'PDF Name'
    sheet['D1'] = 'Page Number'
 
    max_row += 1
    sheet[f'A{max_row}'] = data['Question']
    sheet[f'B{max_row}'] = data['Answer']
    sheet[f'C{max_row}'] = data['PDF Name']
    sheet[f'D{max_row}'] = data['Page Number']
 
    # Save the workbook
    workbook.save(excel_file)
 
    print(f'Data saved in "{excel_file}" successfully.')
 
def main():
    if not os.listdir(input_files_path):
        print(RED+"The directory '{input_files_path}' is empty. Please add PDF files to this folder. Exitting ..."+RESET)
        sys.exit()
    
    else:
        print(BLUE+f"The following PDF files will be processed:"+RESET)
        # List all files in the directory
        files = os.listdir(input_files_path)
        pdf_files = [file for file in files if file.lower().endswith('.pdf')]
        if pdf_files:
            for pdf_file in pdf_files:
                print(f"- {pdf_file}")
        else:
            print(RED+"Please put the pdf files in the "+input_files_path+" . Exitting ..."+RESET)
        # Loop through all files in the input directory 
    
    for filename in os.listdir(input_files_path):
        # Check if the file is a PDF
        if filename.lower().endswith('.pdf'):
            pdf_path = os.path.join(input_files_path, filename)
            # Call the extract function and store the results
            extract_all_from_pdf(pdf_path, output_images_path)
 
    ##prompt
 
    while True:
        question = input(GREEN+"Enter your question (type !exit): "+RESET)
        if question.strip() == "!exit":
            print("Session ended.")
            break
        else:
            (response, relevant_pdf, relevant_page) = ask_question(question)
            answer = response.choices[0].message.content
            data = {
                "Question": question,
                "Answer": answer,
                "PDF Name": relevant_pdf,
                "Page Number": relevant_page
            }
            print(f"Response: {answer}")
            save_qa_to_excel(data)
    
if __name__ == '__main__':
    main()
